"""
batch.py — Batch write, xlsx preview, month-aware file download.
"""
import zipfile
from io import BytesIO
from pathlib import Path
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse, HTMLResponse
import json
from openpyxl import load_workbook

from services.batch_write_svc import (
    batch_write_cht_nokia, batch_write_cht_dk,
    batch_write_wipro, batch_write_project_f, batch_write_nokia_cost,
)
from services.scan_svc import _emp_dir
from services.people_svc import get_all, find_by_name
from config import DEPT_DIR

router = APIRouter()


# ── Batch Write ──────────────────────────────────────────────────

@router.post("/batch-write/cht-nokia")
async def bw_cht_nokia(
    period:     str = Form(""),
    forms_json: str = Form(...),   # {emp_en: {workdays, ess, ot, ta, leave}}
):
    """Write all CHT Nokia employees using fixed server templates."""
    forms = json.loads(forms_json)
    try:
        zip_bytes = batch_write_cht_nokia(period, forms)
    except Exception as e:
        raise HTTPException(500, str(e))
    fname = f"Nokia_工作天數表_{period or 'batch'}.zip"
    return StreamingResponse(
        BytesIO(zip_bytes), media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("/batch-write/cht-dk")
async def bw_cht_dk(
    period:     str = Form(""),
    forms_json: str = Form(...),
):
    forms = json.loads(forms_json)
    try:
        result = batch_write_cht_dk(period, forms)
    except FileNotFoundError as e:
        raise HTTPException(404, str(e))
    fname = f"MN_CHT_工時紀錄表_{period or 'batch'}.xlsx"
    return StreamingResponse(
        BytesIO(result),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("/batch-write/wipro")
async def bw_wipro(
    period:     str = Form(""),
    sheet_name: str = Form(""),
    forms_json: str = Form(...),
    template:   UploadFile = File(...),
):
    forms = json.loads(forms_json)
    tb = await template.read()
    result = batch_write_wipro(tb, sheet_name or period, forms)
    fname = f"SNDA_Dashboard_{period or 'batch'}.xlsx"
    return StreamingResponse(
        BytesIO(result),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("/batch-write/project-f")
async def bw_project_f(
    period:     str = Form(""),
    forms_json: str = Form(...),
    template:   UploadFile = File(...),
):
    forms = json.loads(forms_json)
    tb = await template.read()
    result = batch_write_project_f(tb, forms)
    fname = f"Project_F_{period or 'batch'}.xlsx"
    return StreamingResponse(
        BytesIO(result),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("/batch-write/nokia-cost")
async def bw_nokia_cost(
    period:     str = Form(""),
    sheet_name: str = Form(""),
    forms_json: str = Form(...),
    template:   UploadFile = File(...),
):
    forms = json.loads(forms_json)
    tb = await template.read()
    result = batch_write_nokia_cost(tb, sheet_name or period, forms)
    fname = f"Nokia_費用統整_{period or 'batch'}.xlsx"
    return StreamingResponse(
        BytesIO(result),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── xlsx Preview ─────────────────────────────────────────────────

@router.get("/preview-xlsx/{emp_en}/{file_path:path}")
def preview_xlsx(emp_en: str, file_path: str):
    """Parse xlsx and return as HTML table for frontend rendering."""
    person = find_by_name(emp_en)
    if not person:
        raise HTTPException(404, "Employee not found")
    path = _emp_dir(person) / file_path
    if not path.exists():
        raise HTTPException(404, "File not found")

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return HTMLResponse("<p>空白檔案</p>")

        html = '<table style="border-collapse:collapse;font-size:11px;white-space:nowrap">'
        for i, row in enumerate(rows[:50]):   # limit to 50 rows
            tag = "th" if i == 0 else "td"
            style = "background:#D6E8F8;font-weight:500;" if i == 0 else ""
            html += "<tr>" + "".join(
                f'<{tag} style="border:1px solid #DDE8F6;padding:3px 8px;{style}">'
                f'{("" if v is None else str(v))}</{tag}>'
                for v in row
            ) + "</tr>"
        html += "</table>"
        if len(rows) > 50:
            html += f'<p style="font-size:11px;color:#888;margin-top:6px">只顯示前 50 列（共 {len(rows)} 列）</p>'
        return HTMLResponse(html)
    except Exception as e:
        raise HTTPException(500, f"無法解析 xlsx：{e}")


# ── Download with filter ─────────────────────────────────────────

@router.get("/download-filtered")
def download_filtered(
    q:      str = "",       # name or unit search
    period: str = "",       # e.g. 2026-P05
    ids:    str = "",       # comma-separated employee ids
):
    """Download filtered employees' files as zip."""
    people = get_all()

    # Filter by ids if given
    if ids:
        id_set = {int(x) for x in ids.split(",") if x.strip().isdigit()}
        people = [p for p in people if p["id"] in id_set]
    elif q:
        ql = q.lower()
        people = [p for p in people if
                  ql in (p.get("cn","") + p.get("en","")).lower() or
                  ql in p.get("unit","").lower()]

    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in people:
            emp_dir = _emp_dir(p)
            if not emp_dir.exists():
                continue
            search_root = emp_dir / period if (period and (emp_dir/period).exists()) else emp_dir
            for f in sorted(search_root.rglob("*")):
                if f.is_file():
                    arc = f.relative_to(DEPT_DIR)
                    zf.write(f, arc)

    buf.seek(0)
    fname = f"download_{period or 'all'}.zip"
    return StreamingResponse(
        buf, media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── Get available periods ────────────────────────────────────────

@router.get("/periods")
def get_periods():
    """Return all period folders found across all employees."""
    periods = set()
    if DEPT_DIR.exists():
        for p in DEPT_DIR.rglob("*"):
            if p.is_dir():
                name = p.name
                if name.startswith("20") and "-P" in name:
                    periods.add(name)
    return {"periods": sorted(periods)}
