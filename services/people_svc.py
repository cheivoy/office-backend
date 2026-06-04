import json
from pathlib import Path
from openpyxl import load_workbook
from config import PEOPLE_CSV
from models.schemas import Person

# ── Storage helpers ────────────────────────────────────────────────────────────

def _load() -> list[dict]:
    if PEOPLE_CSV.exists():
        return json.loads(PEOPLE_CSV.read_text(encoding="utf-8"))
    return []

def _save(people: list[dict]):
    PEOPLE_CSV.write_text(json.dumps(people, ensure_ascii=False, indent=2), encoding="utf-8")

# ── Period roster storage ──────────────────────────────────────────────────────
# Each period has its OWN snapshot of the roster, stored in data/roster_<period>.json
# This allows different headcounts per period (joins/departures).

def _roster_path(period: str) -> Path:
    from config import DATA_DIR
    return DATA_DIR / f"roster_{period}.json"

def _load_roster(period: str) -> list[dict] | None:
    """Load period-specific roster. Returns None if not set (fall back to global)."""
    p = _roster_path(period)
    if p.exists():
        return json.loads(p.read_text(encoding="utf-8"))
    return None

def _save_roster(period: str, people: list[dict]):
    _roster_path(period).write_text(
        json.dumps(people, ensure_ascii=False, indent=2), encoding="utf-8"
    )

def get_roster_periods() -> list[str]:
    """Return list of periods that have a saved roster snapshot."""
    from config import DATA_DIR
    periods = []
    for f in sorted(DATA_DIR.glob("roster_*.json")):
        period = f.stem.replace("roster_", "")
        periods.append(period)
    return periods

# ── CRUD (global people list) ──────────────────────────────────────────────────

def get_all(period: str = "") -> list[dict]:
    """Return people for a given period (falls back to global list)."""
    if period:
        roster = _load_roster(period)
        if roster is not None:
            return roster
    return _load()

def upsert(person: Person, period: str = "") -> dict:
    """新增/更新員工。period 有值且該月份有專屬名單時操作該名單，否則操作全域名單。"""
    if period and _load_roster(period) is not None:
        people = _load_roster(period)
        save = lambda lst: _save_roster(period, lst)
    else:
        people = _load()
        save = _save
    if person.id is not None:
        for i, p in enumerate(people):
            if p["id"] == person.id:
                people[i] = person.model_dump()
                save(people)
                return people[i]
    global_max = max((p["id"] for p in _load()), default=0)
    here_max = max((p["id"] for p in people), default=0)
    new_id = max(global_max, here_max) + 1
    data = person.model_dump()
    data["id"] = new_id
    people.append(data)
    save(people)
    return data

def delete(person_id: int, period: str = "") -> bool:
    """刪除員工。period 有值且該月份有專屬名單時只從該名單刪除。"""
    if period and _load_roster(period) is not None:
        return delete_from_roster(period, person_id)
    people = _load()
    new = [p for p in people if p["id"] != person_id]
    if len(new) == len(people):
        return False
    _save(new)
    return True

def delete_from_roster(period: str, person_id: int) -> bool:
    """從某月份專屬名單刪除一個人（不影響全域名單與其他月份）。"""
    roster = _load_roster(period)
    if roster is None:
        return False
    new = [p for p in roster if p.get("id") != person_id]
    if len(new) == len(roster):
        return False
    _save_roster(period, new)
    return True


def delete_roster(period: str) -> bool:
    """整份刪除某月份名單（之後該月份會回退使用全域名單）。"""
    p = _roster_path(period)
    if p.exists():
        p.unlink()
        return True
    return False


def add_to_roster(period: str, person: dict) -> dict:
    """在某月份名單新增一個人（給定 proj/unit/pm/cn/en）。"""
    roster = _load_roster(period)
    if roster is None:
        roster = []
    # 用全域 + 本名單的最大 id 推下一個 id，避免衝突
    global_max = max((p["id"] for p in _load()), default=0)
    roster_max = max((p.get("id", 0) for p in roster), default=0)
    new_id = max(global_max, roster_max) + 1
    entry = {
        "id":   new_id,
        "proj": person.get("proj", ""),
        "unit": person.get("unit", ""),
        "pm":   person.get("pm", ""),
        "cn":   person.get("cn", ""),
        "en":   person.get("en", ""),
    }
    roster.append(entry)
    _save_roster(period, roster)
    return entry


def import_from_excel(file_bytes: bytes, period: str = "") -> list[dict]:
    """
    Read Excel with columns: 專案 / 單位 / PM / 中文姓名 / 英文姓名

    period 有值時：建立／**完全覆蓋**該月份專屬名單快照，
                  不再污染全域名單（避免離職員工永久殘留在全域）。
    period 為空時：合併進全域名單（依英文名 upsert）。
    """
    from io import BytesIO
    wb = load_workbook(BytesIO(file_bytes), read_only=True)
    ws = wb.active

    header_skipped = False
    imported = []
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        if not row or len(row) < 5 or not row[4]:
            continue
        proj, unit, pm, cn, en = (str(v).strip() if v else "" for v in row[:5])
        # Skip summary/total rows
        if unit.lower() in ("total", "subtotal", "合計", "小計", "sum"):
            continue
        if not en:
            continue
        imported.append({"proj": proj, "unit": unit, "pm": pm, "cn": cn, "en": en})

    if period:
        # 用全域名單的英文名沿用既有 id（讓同一人跨月份 id 穩定），
        # 新人就在名單內部給新 id，但「不」寫回全域。
        global_people = _load()
        existing_by_en = {p["en"]: p for p in global_people}
        used_ids = {p["id"] for p in global_people}
        next_id = (max(used_ids) if used_ids else 0) + 1

        roster = []
        for item in imported:
            if item["en"] in existing_by_en:
                base = dict(existing_by_en[item["en"]])
                base.update({k: item[k] for k in ("proj", "unit", "pm", "cn") if item[k]})
                roster.append(base)
            else:
                while next_id in used_ids:
                    next_id += 1
                used_ids.add(next_id)
                roster.append({"id": next_id, **item})
                next_id += 1

        _save_roster(period, roster)   # 完全覆蓋舊的當月名單
        return roster
    else:
        # Legacy: merge into global
        people = _load()
        existing = {p["en"]: p for p in people}
        max_id = max((p["id"] for p in people), default=0)
        for item in imported:
            if item["en"] in existing:
                existing[item["en"]].update({k: item[k] for k in ("proj", "unit", "pm", "cn") if item[k]})
            else:
                max_id += 1
                existing[item["en"]] = {"id": max_id, **item}
        merged = list(existing.values())
        _save(merged)
        return merged

def find_by_id(person_id: int, period: str = "") -> dict | None:
    for p in get_all(period):
        if p.get("id") == person_id:
            return p
    return None


def find_by_name(name: str, period: str = "") -> dict | None:
    """
    嚴格比對：用於 move-file / assign-manual 等「指定特定員工」的場景。

    舊版用 `any(part in name_l ...)` 的鬆散比對會造成嚴重錯置——
    例如有員工英文名含 "chen"，任何檔名/名稱出現 chen 都會誤判成那個人，
    導致「導入時顯示正確、實際歸檔錯置」。

    這裡只接受三種明確的匹配，並要求唯一：
      1. 傳入值正規化後等於某人的英文全名（去空白/底線）
      2. 傳入值等於某人的中文全名
      3. 英文「名+姓」兩段都完整出現（token 完全相符）
    若有多筆符合（歧義）→ 回傳 None，由呼叫端處理，絕不亂猜。
    """
    people = get_all(period)
    raw = name.strip()
    norm = raw.lower().replace("_", "").replace("-", "").replace(" ", "")

    exact = []
    for p in people:
        en = p.get("en", "").strip()
        cn = p.get("cn", "").strip()
        en_norm = en.lower().replace("_", "").replace("-", "").replace(" ", "")
        # 1. 英文全名正規化完全相等
        if en_norm and en_norm == norm:
            exact.append(p)
            continue
        # 2. 中文全名完全相等
        if cn and cn == raw:
            exact.append(p)

    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        return None  # 歧義，不亂猜

    # 3. 英文「名+姓」兩段 token 都完整出現
    tokens = set(
        raw.lower().replace("_", " ").replace("-", " ").split()
    )
    part_matches = []
    for p in people:
        parts = [x for x in p.get("en", "").lower().split() if len(x) >= 2]
        if len(parts) >= 2 and all(part in tokens for part in parts):
            part_matches.append(p)

    if len(part_matches) == 1:
        return part_matches[0]
    return None  # 0 或多筆歧義 → None
