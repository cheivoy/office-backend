"""
report_svc.py — Write to shared report files.

Project F CNS&MN:
  C=專案, D=姓名(英文模糊), E=工作天數, F=ESS, G=差旅費

Nokia 費用統整:
  A=專案, B=單位, C=中文姓名, D=英文姓名
  F=出差日期, G=出差申請金額
  K=夜班日期, L=夜班金額
  N=ESS日期, O=ESS核對金額
  Output: new xlsx WITHOUT formatting (raw data only)
"""
from __future__ import annotations
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from services.name_utils import name_match


# ── Project F ────────────────────────────────────────────────────

def write_project_f(
    template_bytes: bytes,
    emp_en: str, emp_cn: str,
    proj: str,                # "CNS" or "MN"
    work_days: float | None,
    ess_total: float | None,
    ta_total: float | None,
) -> bytes:
    """
    Find employee row by matching proj (col C) AND fuzzy name (col D).
    Write E=work_days, F=ess_total, G=ta_total.
    Return modified file bytes.
    """
    wb = load_workbook(BytesIO(template_bytes))
    ws = wb.active

    PROJ_COL, NAME_COL = 3, 4          # C=3, D=4 (1-based)
    WD_COL, ESS_COL, TA_COL = 5, 6, 7  # E, F, G

    row_idx = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell_proj = str(row[PROJ_COL - 1].value or "").strip().upper()
        cell_name = row[NAME_COL - 1].value
        if cell_proj != proj.upper():
            continue
        if name_match(cell_name, emp_en, emp_cn):
            row_idx = row[0].row
            break

    if row_idx is None:
        raise ValueError(f"Employee '{emp_en}' (proj={proj}) not found in Project F template")

    def _w(col: int, val):
        if val is not None and val != 0:
            ws.cell(row=row_idx, column=col).value = val

    _w(WD_COL, work_days)
    _w(ESS_COL, ess_total)
    _w(TA_COL, ta_total)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Nokia 費用統整 ────────────────────────────────────────────────

def write_nokia_report(
    template_bytes: bytes,
    emp_en: str, emp_cn: str,
    ta_dates: str | None,       # e.g. "0420-0508"
    ta_amount: float | None,
    ns_dates: str | None,       # e.g. "0429  0513"
    ns_amount: float | None,
    ess_dates: str | None,      # e.g. "0501-0531"
    ess_amount: float | None,
    sheet_name: str | None = None,
) -> bytes:
    """
    Find employee row by matching English or Chinese name (col D or C).
    Write: F=ta_dates, G=ta_amount, K=ns_dates, L=ns_amount,
           N=ess_dates, O=ess_amount.
    Return new xlsx WITHOUT cell formatting (raw data only).
    """
    wb_src = load_workbook(BytesIO(template_bytes), data_only=True)

    # choose sheet
    if sheet_name and sheet_name in wb_src.sheetnames:
        ws_src = wb_src[sheet_name]
    elif "5月費用申請" in wb_src.sheetnames:
        ws_src = wb_src["5月費用申請"]
    else:
        ws_src = wb_src.active

    # col indices (1-based): C=3 CN, D=4 EN, F=6 ta_date, G=7 ta_amt,
    # K=11 ns_date, L=12 ns_amt, N=14 ess_date, O=15 ess_amt
    CN_COL, EN_COL = 3, 4
    TA_D, TA_A = 6, 7
    NS_D, NS_A = 11, 12
    ESS_D, ESS_A = 14, 15

    row_idx = None
    for row in ws_src.iter_rows(min_row=2, values_only=False):
        cn_val = row[CN_COL - 1].value
        en_val = row[EN_COL - 1].value
        if name_match(en_val, emp_en, emp_cn) or name_match(cn_val, emp_en, emp_cn):
            row_idx = row[0].row
            break

    if row_idx is None:
        raise ValueError(f"Employee '{emp_en}' not found in Nokia 費用統整 template")

    # Write to source workbook (preserve row structure, strip formats on output)
    def _w(col: int, val):
        if val is not None and val != "" and val != 0:
            ws_src.cell(row=row_idx, column=col).value = val

    _w(TA_D,  ta_dates)
    _w(TA_A,  ta_amount)
    _w(NS_D,  ns_dates)
    _w(NS_A,  ns_amount)
    _w(ESS_D, ess_dates)
    _w(ESS_A, ess_amount)

    # Build new workbook WITHOUT formatting
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = ws_src.title

    # Copy header row
    headers = [cell.value for cell in ws_src[1]]
    ws_out.append(headers)
    ws_out[1][0].font = Font(bold=True)

    # Copy all data rows (values only)
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            ws_out.append(list(row))

    out = BytesIO()
    wb_out.save(out)
    return out.getvalue()
