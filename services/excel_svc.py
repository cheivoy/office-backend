from io import BytesIO
from openpyxl import load_workbook
from models.schemas import SubmitPayload, OtEntry, LeaveEntry
from config import CHT_COL, WIPRO_COL


# ── Format helpers ────────────────────────────────────────────────

def _fmt_ot(entries: list[OtEntry]) -> str:
    parts = []
    for e in entries:
        if not (e.date and e.tstart and e.tend):
            continue
        mm = e.date[5:].replace("-", "")          # "2026-05-23" → "0523"
        ts = e.tstart.replace(":", "")             # "09:00" → "0900"
        te = e.tend.replace(":", "")
        hrs = f"_{e.hours}hrs" if e.hours else ""
        parts.append(f"{mm}_{ts}-{te}{hrs}")
    return "  ".join(parts)


def _fmt_leave(entries: list[LeaveEntry]) -> str:
    """
    Format leave entries with 8-hour rule:
    - hours >= 8 (full day): "0504, 0514_sick leave"
    - hours < 8 (partial):   "0504, 0514_3hrs sick leave"
    """
    from datetime import date, timedelta
    def expand_range(from_date: str, to_date: str) -> list[str]:
        if not from_date:
            return []
        if not to_date or to_date == from_date:
            return [from_date[5:].replace("-", "")]
        cur = date.fromisoformat(from_date)
        end = date.fromisoformat(to_date)
        dates = []
        while cur <= end:
            dates.append(cur.strftime("%m%d"))
            cur += timedelta(days=1)
        return dates

    parts = []
    for e in entries:
        if not e.from_date:
            continue
        dates = expand_range(e.from_date, e.to_date or e.from_date)
        ds = ", ".join(dates)
        label = e.reason if (e.type == "other" and e.reason) else e.type

        # Calculate hours from time range
        hrs_val = None
        if e.tstart and e.tend:
            sh, sm = map(int, e.tstart.split(":"))
            eh, em = map(int, e.tend.split(":"))
            mins = (eh * 60 + em) - (sh * 60 + sm)
            if mins < 0:
                mins += 1440
            hrs_val = mins / 60

        if hrs_val is not None:
            if hrs_val >= 8:
                hrs_part = "_"          # full day, no time suffix
            else:
                h = int(hrs_val) if hrs_val == int(hrs_val) else round(hrs_val, 1)
                hrs_part = f"_{h}hrs "
        elif e.hours:
            hrs_part = f"_{e.hours} "
        else:
            hrs_part = "_"

        parts.append(f"{ds}{hrs_part}{label}")
    return "  ".join(parts)


def _fmt_ot_wipro(entries: list[OtEntry]) -> str:
    """Wipro format: 0402_1830-2230_4hrs"""
    parts = []
    for e in entries:
        if not (e.date and e.tstart and e.tend):
            continue
        mm = e.date[5:].replace("-", "")
        ts = e.tstart.replace(":", "")
        te = e.tend.replace(":", "")
        hrs = f"_{e.hours}hrs" if e.hours else ""
        parts.append(f"{mm}_{ts}-{te}{hrs}")
    return "  ".join(parts)


# ── Find employee row ─────────────────────────────────────────────

def _find_row(ws, emp_en: str, emp_cn: str, name_col: int = 2) -> int | None:
    """
    Search for the employee row in the worksheet.
    name_col is 0-based; openpyxl uses 1-based so we add 1.
    Matches on partial English name or Chinese name.
    """
    col_idx = name_col + 1
    en_l = emp_en.lower()
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell = row[name_col]
        val = str(cell.value or "").lower()
        if en_l in val or (emp_cn and emp_cn in str(cell.value or "")):
            return cell.row
    return None


# ── CHT Nokia / CHT DK ───────────────────────────────────────────

def write_cht(template_bytes: bytes, payload: SubmitPayload) -> bytes:
    """
    Load the uploaded CHT template, find the employee row,
    write only the requested fields, return modified bytes.
    All existing styles / formulas / other cells are untouched.
    """
    wb = load_workbook(BytesIO(template_bytes))
    ws = wb.active

    row_idx = _find_row(ws, payload.emp_en, payload.emp_name)
    if row_idx is None:
        raise ValueError(f"Employee '{payload.emp_en}' not found in template")

    def _write(col_key: str, value):
        if value is None or value == "" or value == 0:
            return
        col = CHT_COL[col_key] + 1   # openpyxl is 1-based
        ws.cell(row=row_idx, column=col).value = value

    if payload.work_days is not None:
        _write("work_days", payload.work_days)

    if payload.leave:
        _write("leave", _fmt_leave(payload.leave))

    if payload.ot:
        _write("ot", _fmt_ot(payload.ot))

    ess_total = sum(e.amount or 0 for e in payload.ess)
    ns_total  = sum(e.ns_amount or 0 for e in payload.ess)
    ta_total  = sum(t.amount or 0 for t in payload.ta)

    if ess_total:
        _write("ess", ess_total)
    if ns_total:
        _write("night_shift", ns_total)
    if ta_total:
        _write("travel", ta_total)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Wipro SNDA ───────────────────────────────────────────────────

def write_wipro(template_bytes: bytes, emp_en: str,
                ess: float = 0, shift: float = 0,
                ot_entries: list[OtEntry] = None,
                travel: float = 0,
                sheet_name: str = None) -> bytes:
    """
    Write to Wipro SNDA Dashboard.
    sheet_name: e.g. 'P06_26'. If None, uses the last sheet.
    Creates the sheet if it doesn't exist yet.
    """
    wb = load_workbook(BytesIO(template_bytes))

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            # Copy structure from last data sheet
            src = wb.worksheets[-1]
            ws = wb.copy_worksheet(src)
            ws.title = sheet_name
        else:
            ws = wb[sheet_name]
    else:
        ws = wb.worksheets[-1]

    row_idx = _find_row(ws, emp_en, "", name_col=2)
    if row_idx is None:
        raise ValueError(f"Employee '{emp_en}' not found in sheet '{ws.title}'")

    def _w(col_key: str, value):
        if value is None or value == "" or value == 0:
            return
        col = WIPRO_COL[col_key] + 1
        ws.cell(row=row_idx, column=col).value = value

    if ess:
        _w("ess", ess)
    if shift:
        _w("shift", shift)
    if ot_entries:
        _w("ot", _fmt_ot_wipro(ot_entries))
    if travel:
        _w("travel", travel)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
