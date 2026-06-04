"""
verify_svc.py — Wipro SNDA cross-check service.

All functions accept:
  - template_bytes: bytes of the uploaded approval/record file
  - emp_en / emp_cn: employee names from people.json
  - tab_data: data from the fill-in form (TAB)

Return a VerifyResult with status, matched records, and anomalies.
"""
from __future__ import annotations
from io import BytesIO
from dataclasses import dataclass, field
from openpyxl import load_workbook
from services.name_utils import (
    name_match, parse_date, parse_mmdd, parse_time_range, hours_from_range
)


@dataclass
class Anomaly:
    field: str
    expected: str
    found: str
    note: str = ""


def _expand_dates(entry: dict) -> list[str]:
    """Return list of YYYY-MM-DD dates an entry covers.
    Supports new range form (from_date/to_date) and legacy single `date`."""
    from datetime import date, timedelta
    fd = entry.get("from_date") or entry.get("date") or ""
    td = entry.get("to_date") or fd
    fd, td = parse_date(fd), parse_date(td)
    if not fd:
        return []
    if not td or td < fd:
        td = fd
    try:
        y1, m1, d1 = map(int, fd.split("-"))
        y2, m2, d2 = map(int, td.split("-"))
        cur, end = date(y1, m1, d1), date(y2, m2, d2)
    except (ValueError, AttributeError):
        return [fd]
    out = []
    while cur <= end and len(out) < 60:
        out.append(cur.isoformat())
        cur += timedelta(days=1)
    return out


@dataclass
class VerifyResult:
    emp_en: str
    emp_cn: str
    status: str           # "ok" | "anomaly" | "not_found"
    matched_rows: int = 0
    anomalies: list[Anomaly] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "emp": self.emp_en,
            "status": self.status,
            "matched_rows": self.matched_rows,
            "anomalies": [a.__dict__ for a in self.anomalies],
        }


# ── helpers ────────────────────────────────────────────────────────

def _iter_approved(ws, name_col: int, status_col: int,
                   emp_en: str, emp_cn: str):
    """Yield rows where name matches AND status is Approve/Approved."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        cell_name = row[name_col] if len(row) > name_col else None
        status = str(row[status_col] or "").strip().lower() if len(row) > status_col else ""
        if not name_match(cell_name, emp_en, emp_cn):
            continue
        if "approv" not in status:
            continue
        yield row


def _dedup(rows: list[tuple], key_fn) -> list[tuple]:
    """De-duplicate rows with identical keys."""
    seen: set = set()
    result = []
    for row in rows:
        k = key_fn(row)
        if k not in seen:
            seen.add(k)
            result.append(row)
    return result


# ── Travel ────────────────────────────────────────────────────────

def verify_travel(
    template_bytes: bytes,
    emp_en: str, emp_cn: str,
    tab_ta: list[dict],          # [{from_date, to_date, amount}]
) -> VerifyResult:
    """
    Travel DataSheet:
      B(1)=Traveler Name, F(5)=Status, K(10)=Start, L(11)=End, AC(28)=Total expenses
    """
    wb = load_workbook(BytesIO(template_bytes), read_only=True, data_only=True)
    ws = wb.active
    result = VerifyResult(emp_en=emp_en, emp_cn=emp_cn, status="not_found")

    NAME_COL, STATUS_COL, START_COL, END_COL, TOTAL_COL = 1, 17, 10, 11, 28

    raw_rows = list(_iter_approved(ws, NAME_COL, STATUS_COL, emp_en, emp_cn))
    rows = _dedup(raw_rows, lambda r: (parse_date(r[START_COL]), parse_date(r[END_COL])))
    if not rows:
        return result

    result.matched_rows = len(rows)
    result.status = "ok"

    for tab in tab_ta:
        tab_from = tab.get("from_date", "")
        tab_to   = tab.get("to_date", "")
        tab_amt  = float(tab.get("amount") or 0)

        # find matching row by date range
        match = None
        for row in rows:
            if parse_date(row[START_COL]) == tab_from and parse_date(row[END_COL]) == tab_to:
                match = row
                break

        if match is None:
            result.status = "anomaly"
            result.anomalies.append(Anomaly(
                field="travel_date",
                expected=f"{tab_from} ~ {tab_to}",
                found="查無對應紀錄",
                note="Approval 中找不到相符的差旅日期區間",
            ))
            continue

        approved_total = float(match[TOTAL_COL] or 0)
        if tab_amt > approved_total:
            result.status = "anomaly"
            result.anomalies.append(Anomaly(
                field="travel_amount",
                expected=f"≤ NT${approved_total:,.0f}",
                found=f"NT${tab_amt:,.0f}",
                note=f"申請金額超過 Approval Total Expenses（差額 NT${tab_amt - approved_total:,.0f}）",
            ))

    return result


# ── OT ────────────────────────────────────────────────────────────

def verify_ot(
    template_bytes: bytes,
    emp_en: str, emp_cn: str,
    tab_ot: list[dict],          # [{date, tstart, tend, hours}]
) -> VerifyResult:
    """
    OT Data sheet:
      B(1)=Name, F(5)=Response, R(17)=OT Date, S(18)=time range, U(20)=Final Total Hrs
    """
    wb = load_workbook(BytesIO(template_bytes), read_only=True, data_only=True)
    ws = wb["OT Data"]
    result = VerifyResult(emp_en=emp_en, emp_cn=emp_cn, status="not_found")

    NAME_COL, STATUS_COL = 1, 5
    DATE_COL, TIME_COL, HRS_COL = 17, 18, 20

    raw_rows = list(_iter_approved(ws, NAME_COL, STATUS_COL, emp_en, emp_cn))
    rows = _dedup(raw_rows, lambda r: (parse_date(r[DATE_COL]), str(r[TIME_COL] or "").strip()))
    if not rows:
        return result

    result.matched_rows = len(rows)
    result.status = "ok"

    for tab in tab_ot:
        tab_date  = tab.get("date", "")
        tab_start = tab.get("tstart", "")
        tab_end   = tab.get("tend", "")
        tab_hrs   = float(tab.get("hours") or 0)

        match = None
        for row in rows:
            row_date = parse_date(row[DATE_COL])
            row_time = parse_time_range(str(row[TIME_COL] or ""))
            if row_date != tab_date:
                continue
            if row_time and row_time[0] == tab_start and row_time[1] == tab_end:
                match = row
                break
            # relax: date match only if time not parseable
            if row_time is None:
                match = row
                break

        if match is None:
            result.status = "anomaly"
            result.anomalies.append(Anomaly(
                field="ot_record",
                expected=f"{tab_date} {tab_start}-{tab_end}",
                found="查無對應紀錄",
                note="OT Data 中找不到相符的日期/時間",
            ))
            continue

        # check hours
        approved_hrs = float(match[HRS_COL] or 0)
        if tab_hrs and abs(tab_hrs - approved_hrs) > 0.1:
            result.status = "anomaly"
            result.anomalies.append(Anomaly(
                field="ot_hours",
                expected=f"{approved_hrs}hrs",
                found=f"{tab_hrs}hrs",
                note=f"填寫時數與 Approval 不符（{tab_date}）",
            ))

    return result


# ── Night Shift ───────────────────────────────────────────────────

def verify_ns(
    template_bytes: bytes,
    emp_en: str, emp_cn: str,
    tab_ess: list[dict],          # ESS entries that have ns_amount > 0, reuse date/tstart/tend
) -> VerifyResult:
    """
    OT_Shift_01June26 sheet:
      B(1)=Name, F(5)=Status, Q(16)=Shift Date, R(17)=time range
    """
    wb = load_workbook(BytesIO(template_bytes), read_only=True, data_only=True)
    ws = wb["OT_Shift_01June26"]
    result = VerifyResult(emp_en=emp_en, emp_cn=emp_cn, status="not_found")

    NAME_COL, STATUS_COL = 1, 5
    DATE_COL, TIME_COL = 16, 17

    raw_rows = list(_iter_approved(ws, NAME_COL, STATUS_COL, emp_en, emp_cn))
    rows = _dedup(raw_rows, lambda r: (parse_date(r[DATE_COL]), str(r[TIME_COL] or "").strip()))
    if not rows:
        return result

    result.matched_rows = len(rows)
    result.status = "ok"

    # NS entries now come from their own array. Support both new (ns[].amount,
    # any positive) and legacy (ess[].ns_amount > 0) shapes.
    ns_entries = [e for e in tab_ess
                  if float(e.get("amount") or 0) > 0 or float(e.get("ns_amount") or 0) > 0]

    for tab in ns_entries:
        tab_date  = tab.get("date", "")
        tab_start = tab.get("tstart", "")
        tab_end   = tab.get("tend", "")

        match = None
        for row in rows:
            if parse_date(row[DATE_COL]) != tab_date:
                continue
            row_time = parse_time_range(str(row[TIME_COL] or ""))
            if row_time and row_time[0] == tab_start and row_time[1] == tab_end:
                match = row
                break
            if row_time is None:
                match = row
                break

        if match is None:
            result.status = "anomaly"
            result.anomalies.append(Anomaly(
                field="ns_record",
                expected=f"{tab_date} {tab_start}-{tab_end}",
                found="查無對應 Night Shift 紀錄",
                note="OT_Shift sheet 中找不到相符的日期/時間",
            ))

    return result


# ── ESS ROTA ─────────────────────────────────────────────────────

def verify_ess(
    template_bytes: bytes,
    emp_en: str, emp_cn: str,
    tab_ess: list[dict],          # [{date, amount}] — one entry per ESS day
    tab_ess_total: float,
) -> VerifyResult:
    """
    ESS ROTA「明細」sheet (header row 2, data from row 3):
      B(1)=人員, D(3)=日期, H(7)=費率
    Extract all approved ESS dates for this person, sum amounts,
    compare with tab total.
    """
    wb = load_workbook(BytesIO(template_bytes), read_only=True, data_only=True)
    ws = wb["明細"]
    result = VerifyResult(emp_en=emp_en, emp_cn=emp_cn, status="not_found")

    NAME_COL, DATE_COL, AMT_COL = 1, 3, 7

    # Collect all date→amount for this person (明細 has one row per day)
    rota: dict[str, float] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        cell_name = row[NAME_COL] if len(row) > NAME_COL else None
        if not name_match(cell_name, emp_en, emp_cn):
            continue
        d = parse_date(row[DATE_COL] if len(row) > DATE_COL else None)
        amt = float(row[AMT_COL] if len(row) > AMT_COL and row[AMT_COL] else 0)
        if d:
            rota[d] = rota.get(d, 0) + amt

    if not rota:
        return result

    result.matched_rows = len(rota)
    result.status = "ok"

    # Expand each ESS entry's date range into individual days
    tab_dates = set()
    for e in tab_ess:
        for d in _expand_dates(e):
            tab_dates.add(d)

    # Check each ESS date exists in ROTA
    for tab_date in sorted(tab_dates):
        if tab_date not in rota:
            result.status = "anomaly"
            result.anomalies.append(Anomaly(
                field="ess_date",
                expected=tab_date,
                found="不在 ROTA 排班表中",
                note=f"{tab_date} 未出現在 ESS ROTA 明細，請確認排班",
            ))

    # Sum amount for TAB dates and compare with stated total
    rota_sum = sum(v for k, v in rota.items() if k in tab_dates)

    if tab_ess_total and abs(rota_sum - tab_ess_total) > 1:
        result.status = "anomaly"
        result.anomalies.append(Anomaly(
            field="ess_amount",
            expected=f"NT${tab_ess_total:,.0f}（填寫值）",
            found=f"NT${rota_sum:,.0f}（ROTA 計算值）",
            note=f"金額差異 NT${abs(rota_sum - tab_ess_total):,.0f}，請核對費率（平日 100／假日 500）",
        ))

    return result